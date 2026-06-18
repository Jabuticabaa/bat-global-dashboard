import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from pathlib import Path
from openpyxl import load_workbook

BASE = Path(r"C:\Users\Marcelo - 1326\BP One Group\Bloco TSC - Documentos\General\Clientes\Atuais\BAT Global\Estudos\Expansão mercados\Arquivos base")
F2 = BASE / "BAT_Global_Market_Intelligence ATT.xlsx"

wb = load_workbook(F2, read_only=True, data_only=True)
print("Abas disponíveis:", wb.sheetnames)

# Inspecionar By Region completo
ws = wb["📍 By Region"]
rows = list(ws.iter_rows(min_row=4, values_only=True))
hdr = [str(c) if c else "" for c in rows[0]]
print("\nHeader By Region:", hdr[:10])

# Contar regiões por país
by_country: dict = {}
for row in rows[1:]:
    if not row[1] or not row[2] or not row[3]: continue
    country = str(row[2]).strip()
    region  = str(row[3]).strip()
    if country not in by_country:
        by_country[country] = []
    by_country[country].append(region)

print(f"\nTotal países com regiões: {len(by_country)}")
print("\nPaíses e suas regiões:")
for country, regions in sorted(by_country.items()):
    print(f"  {country} ({len(regions)}): {regions[:5]}{'...' if len(regions)>5 else ''}")

# Verificar aba By Country - quais colunas
ws2 = wb["🌍 By Country"]
rows2 = list(ws2.iter_rows(min_row=4, values_only=True))
hdr2 = [str(c) if c else "" for c in rows2[0]]
print("\nHeader By Country:", hdr2)

wb.close()
