"""
BAT Dashboard — ETL + Correlation Analysis
Reads both Excel files, validates JOIN keys, geocodes locations,
outputs structured JSON hierarchy L0→L5 ready for the map.
"""

import json, re, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

# ─── PATHS ───────────────────────────────────────────────────────────────────
BASE = Path(r"C:\Users\Marcelo - 1326\BP One Group\Bloco TSC - Documentos\General\Clientes\Atuais\BAT Global\Estudos\Expansão mercados\Arquivos base")
F1 = BASE / "Dados BAT Total.xlsx"
F2 = BASE / "BAT_Global_Market_Intelligence ATT.xlsx"
OUT = Path(r"C:\Users\Marcelo - 1326\bat-dashboard\public\data")
OUT.mkdir(parents=True, exist_ok=True)

# ─── L1 CONTINENT GROUPING ───────────────────────────────────────────────────
CONTINENT_MAP = {
    "North America":         "América do Norte",
    "Central America":       "América do Norte",
    "Caribe":                "América do Norte",
    "South America":         "América Latina / Sul",
    "Europe Ocidental":      "Europa",
    "Europe Oriental":       "Europa",
    "Europe do Sul":         "Europa",
    "Africa do Norte":       "África",
    "Africa Ocidental":      "África",
    "Africa Oriental":       "África",
    "Africa Central":        "África",
    "Africa do Sul/Austral": "África",
    "Africa Insular":        "África",
    "Asia do Sul":           "Ásia",
    "Asia Oriental":         "Ásia",
    "Sudeste Asiático":      "Ásia",
    "Oriente Médio":         "Ásia",
    "Oceania":               "Oceania",
}

L1_ORDER = ["América do Norte","América Latina / Sul","Europa","África","Ásia","Oceania"]

# ─── BRAZIL/CHILE REGION → NAME MAPPING (JOIN KEY) ───────────────────────────
# BDOS Descrição(Local) prefix → Market Intelligence Region code
BDOS_REGION_MAP = {
    "NNE": "NNE",
    "CTO": "CTO",
    "RIO": "RIO",
    "SPC": "SPC",
    "SPR": "SPR",
    "SUL": "SUL",
    "CEM": "CTO",   # CEM is a sub-depot of CTO (Uberlândia)
}

# BDOS Nome Ciclo → Market Intelligence Country
CICLO_TO_COUNTRY = {
    "BAT BRASIL": "Brazil",
    "BAT CHILE":  "Chile",
}

# ─── GEOCODING TABLE — Level 5 cities (Brazil operational) ───────────────────
# Standard: WGS84 lat/lng — universal for Leaflet/GeoJSON
GEO_L5 = {
    # NNE ──────────────────────────────────────────────────────────────────────
    "NNE - CID - Recife/PE":                 {"lat":-8.0476,  "lng":-34.8770, "city":"Recife",        "state":"PE","region":"NNE","country":"Brazil"},
    "NNE - COD - Aracaju/SE":                {"lat":-10.9472, "lng":-37.0731, "city":"Aracaju",       "state":"SE","region":"NNE","country":"Brazil"},
    "NNE - COD - Belem/PA":                  {"lat":-1.4558,  "lng":-48.5044, "city":"Belém",         "state":"PA","region":"NNE","country":"Brazil"},
    "NNE - COD - Fortaleza/CE":              {"lat":-3.7327,  "lng":-38.5270, "city":"Fortaleza",     "state":"CE","region":"NNE","country":"Brazil"},
    "NNE - COD - Joao Pessoa/PB":            {"lat":-7.1195,  "lng":-34.8450, "city":"João Pessoa",   "state":"PB","region":"NNE","country":"Brazil"},
    "NNE - COD - MACAPA/AP":                 {"lat":0.0349,   "lng":-51.0694, "city":"Macapá",        "state":"AP","region":"NNE","country":"Brazil"},
    "NNE - COD - Maceio/AL":                 {"lat":-9.6658,  "lng":-35.7350, "city":"Maceió",        "state":"AL","region":"NNE","country":"Brazil"},
    "NNE - COD - Manaus/AM":                 {"lat":-3.1019,  "lng":-60.0250, "city":"Manaus",        "state":"AM","region":"NNE","country":"Brazil"},
    "NNE - COD - Maraba/PA":                 {"lat":-5.3686,  "lng":-49.1178, "city":"Marabá",        "state":"PA","region":"NNE","country":"Brazil"},
    "NNE - COD - PORTO VELHO/RO":            {"lat":-8.7612,  "lng":-63.9004, "city":"Porto Velho",   "state":"RO","region":"NNE","country":"Brazil"},
    "NNE - COD - SAO LUIS /MA":              {"lat":-2.5307,  "lng":-44.3068, "city":"São Luís",      "state":"MA","region":"NNE","country":"Brazil"},
    "NNE - Distribuidora Riograndense - RN": {"lat":-5.7945,  "lng":-35.2110, "city":"Natal",         "state":"RN","region":"NNE","country":"Brazil"},
    "NNE - GD Distribudor - PI":             {"lat":-5.0892,  "lng":-42.8019, "city":"Teresina",      "state":"PI","region":"NNE","country":"Brazil"},
    "NNE - ISD DISTRIBUIDORA RECOL - AC":    {"lat":-9.9754,  "lng":-67.8249, "city":"Rio Branco",    "state":"AC","region":"NNE","country":"Brazil"},
    # CTO ──────────────────────────────────────────────────────────────────────
    "CTO - CID - Contagem/MG":               {"lat":-19.9317, "lng":-44.0536, "city":"Contagem",      "state":"MG","region":"CTO","country":"Brazil"},
    "CTO - COD - Brasilia/DF":               {"lat":-15.7801, "lng":-47.9292, "city":"Brasília",      "state":"DF","region":"CTO","country":"Brazil"},
    "CTO - COD - Campo Grande/MS":           {"lat":-20.4697, "lng":-54.6201, "city":"Campo Grande",  "state":"MS","region":"CTO","country":"Brazil"},
    "CTO - COD - Cuiaba/MT":                 {"lat":-15.6014, "lng":-56.0979, "city":"Cuiabá",        "state":"MT","region":"CTO","country":"Brazil"},
    "CTO - COD - Goiania/GO":                {"lat":-16.6869, "lng":-49.2648, "city":"Goiânia",       "state":"GO","region":"CTO","country":"Brazil"},
    "CTO - COD - Uberlandia/MG":             {"lat":-18.9186, "lng":-48.2772, "city":"Uberlândia",    "state":"MG","region":"CTO","country":"Brazil"},
    "CEM - UBERLANDIA/MG":                   {"lat":-18.9186, "lng":-48.2772, "city":"Uberlândia",    "state":"MG","region":"CTO","country":"Brazil"},
    # RIO ──────────────────────────────────────────────────────────────────────
    "RIO - CID - Rio de Janeiro/RJ":         {"lat":-22.9068, "lng":-43.1729, "city":"Rio de Janeiro","state":"RJ","region":"RIO","country":"Brazil"},
    "RIO - COD - Itabuna/BA":                {"lat":-14.7860, "lng":-39.2799, "city":"Itabuna",       "state":"BA","region":"RIO","country":"Brazil"},
    "RIO - COD - Salvador/BA":               {"lat":-12.9714, "lng":-38.5014, "city":"Salvador",      "state":"BA","region":"RIO","country":"Brazil"},
    "RIO - COD - Vila Velha/ES":             {"lat":-20.3297, "lng":-40.2925, "city":"Vila Velha",    "state":"ES","region":"RIO","country":"Brazil"},
    # SPC ──────────────────────────────────────────────────────────────────────
    "SPC - CID - SAO PAULO/SP":              {"lat":-23.5505, "lng":-46.6333, "city":"São Paulo",     "state":"SP","region":"SPC","country":"Brazil"},
    "SPC - COD - Praia Grande/SP":           {"lat":-24.0059, "lng":-46.4022, "city":"Praia Grande",  "state":"SP","region":"SPC","country":"Brazil"},
    # SPR ──────────────────────────────────────────────────────────────────────
    "SPR  - CID - Curitiba/PR":              {"lat":-25.4284, "lng":-49.2733, "city":"Curitiba",       "state":"PR","region":"SPR","country":"Brazil"},
    "SPR  - COD - Bauru/SP":                 {"lat":-22.3246, "lng":-49.0731, "city":"Bauru",          "state":"SP","region":"SPR","country":"Brazil"},
    "SPR  - COD - Campinas/SP - SPR":        {"lat":-22.9056, "lng":-47.0608, "city":"Campinas",       "state":"SP","region":"SPR","country":"Brazil"},
    "SPR  - COD - Londrina/PR":              {"lat":-23.3045, "lng":-51.1696, "city":"Londrina",        "state":"PR","region":"SPR","country":"Brazil"},
    "SPR  - COD - RIBEIRAO PRETO/SP":        {"lat":-21.1775, "lng":-47.8103, "city":"Ribeirão Preto",  "state":"SP","region":"SPR","country":"Brazil"},
    # SUL ──────────────────────────────────────────────────────────────────────
    "SUL - CID - Porto Alegre/RS":           {"lat":-30.0346, "lng":-51.2177, "city":"Porto Alegre",   "state":"RS","region":"SUL","country":"Brazil"},
    "SUL - COD - Itajai/SC":                 {"lat":-26.9076, "lng":-48.6608, "city":"Itajaí",         "state":"SC","region":"SUL","country":"Brazil"},
    "SUL - COD - Passo Fundo/RS":            {"lat":-28.2620, "lng":-52.4083, "city":"Passo Fundo",    "state":"RS","region":"SUL","country":"Brazil"},
    # NNE extra ────────────────────────────────────────────────────────────────
    "NNE - COD - Boa Vista/RR":              {"lat":-2.8235,  "lng":-60.6758, "city":"Boa Vista",      "state":"RR","region":"NNE","country":"Brazil"},
    "NNE - COD - NATAL/RN":                  {"lat":-5.7945,  "lng":-35.2110, "city":"Natal",          "state":"RN","region":"NNE","country":"Brazil"},
    "NNE - COD - Palmas/TO":                 {"lat":-10.2491, "lng":-48.3243, "city":"Palmas",         "state":"TO","region":"NNE","country":"Brazil"},
    "NNE - COD - Santarem/PA":               {"lat":-2.4418,  "lng":-54.7082, "city":"Santarém",       "state":"PA","region":"NNE","country":"Brazil"},
    "NNE - COD - TERESINA/PI":               {"lat":-5.0892,  "lng":-42.8019, "city":"Teresina",       "state":"PI","region":"NNE","country":"Brazil"},
}

# Chile regions centroids (L4 for Chile — no L5 operational data yet)
GEO_CHILE_REGIONS = {
    "ANTOFAGASTA": {"lat":-23.6509, "lng":-70.3975},
    "ARICA":       {"lat":-18.4783, "lng":-70.3126},
    "CALAMA":      {"lat":-22.4557, "lng":-68.9183},
    "CASTRO":      {"lat":-42.4827, "lng":-73.7626},
    "CHILLAN":     {"lat":-36.6063, "lng":-72.1034},
    "CONCEPCION":  {"lat":-36.8270, "lng":-73.0498},
    "COPIAPO":     {"lat":-27.3668, "lng":-70.3323},
    "COQUIMBO":    {"lat":-29.9533, "lng":-71.3395},
    "IQUIQUE":     {"lat":-20.2133, "lng":-70.1503},
    "LOS ANGELES": {"lat":-37.4710, "lng":-72.3538},
    "OSORNO":      {"lat":-40.5736, "lng":-73.1343},
    "PUERTO MONTT":{"lat":-41.4717, "lng":-72.9369},
    "PUNTA ARENAS":{"lat":-53.1638, "lng":-70.9171},
    "SAN FERNANDO":{"lat":-34.5867, "lng":-70.9910},
    "SANTIAGO":    {"lat":-33.4489, "lng":-70.6693},
    "TALCA":       {"lat":-35.4264, "lng":-71.6554},
    "TEMUCO":      {"lat":-38.7359, "lng":-72.5904},
    "VALDIVIA":    {"lat":-39.8142, "lng":-73.2459},
    "VALPARAISO":  {"lat":-33.0472, "lng":-71.6127},
}

COUNTRY_CENTROIDS: dict[str, tuple[float, float]] = {
    "Africa do Sul": (-29.0, 25.0), "Albânia": (41.1, 20.2), "Alemanha": (51.2, 10.5),
    "Angola": (-11.2, 17.9), "Argentina": (-38.4, -63.6), "Argélia": (28.0, 1.7),
    "Armênia": (40.1, 45.0), "Arábia Saudita": (24.2, 45.1), "Austrália": (-25.3, 133.8),
    "Azerbaijão": (40.1, 47.6), "Bahrain": (26.0, 50.6), "Bangladesh": (23.7, 90.4),
    "Barbados": (13.2, -59.6), "Benin": (9.3, 2.3), "Botswana": (-22.3, 24.7),
    "Brazil": (-14.2, -51.9), "Bulgária": (42.7, 25.5), "Burkina Faso": (12.4, -1.6),
    "Burundi": (-3.4, 30.0), "Bélgica": (50.5, 4.5), "Camarões": (3.9, 11.5),
    "Camboja": (12.6, 104.9), "Canadá": (56.1, -96.3), "Cazaquistão": (48.0, 66.9),
    "Chile": (-35.7, -71.5), "China": (35.9, 104.2), "Chipre": (35.1, 33.4),
    "Colômbia": (4.1, -72.3), "Comores": (-11.9, 43.3), "Coreia do Sul": (35.9, 127.8),
    "Costa Rica": (9.7, -83.8), "Costa do Marfim": (7.5, -5.5), "Dinamarca": (56.3, 9.5),
    "Djibuti": (11.8, 42.6), "EAU": (24.0, 54.0), "EUA": (37.1, -95.7),
    "Egito": (26.8, 30.8), "El Salvador": (13.8, -88.9), "Equador": (-1.8, -78.2),
    "Eritreia": (15.2, 39.8), "Eslováquia": (48.7, 19.7), "Eslovênia": (46.2, 14.8),
    "Espanha": (40.5, -3.7), "Estônia": (58.6, 25.0), "Etiópia": (9.1, 40.5),
    "Fiji": (-17.7, 178.1), "Finlândia": (61.9, 25.7), "França": (46.2, 2.2),
    "Gabão": (-0.8, 11.6), "Gana": (7.9, -1.0), "Geórgia": (42.3, 43.4),
    "Grécia": (38.8, 22.0), "Guatemala": (15.8, -90.2), "Guiana": (5.0, -58.9),
    "Guiné": (11.0, -10.9), "Guiné Equatorial": (1.7, 10.3), "Gâmbia": (13.5, -15.3),
    "Holanda": (52.1, 5.3), "Honduras": (15.2, -86.2), "Hong Kong": (22.4, 114.1),
    "Hungria": (47.2, 19.5), "Ilhas Salomão": (-9.4, 160.2), "Indonésia": (-0.8, 113.9),
    "Iraque": (33.2, 43.7), "Irlanda": (53.4, -8.2), "Itália": (42.8, 12.8),
    "Iêmen": (15.6, 48.5), "Jamaica": (18.1, -77.3), "Japão": (36.2, 138.3),
    "Jordânia": (30.6, 36.2), "Kosovo": (42.6, 20.9), "Kuwait": (29.3, 47.5),
    "Letônia": (56.9, 24.6), "Lituânia": (55.2, 23.9), "Líbano": (33.9, 35.5),
    "Líbia": (26.3, 17.2), "Macedônia": (41.6, 21.7), "Madagascar": (-18.8, 46.9),
    "Malaui": (-13.3, 34.3), "Mali": (17.6, -4.0), "Malta": (35.9, 14.4),
    "Malásia": (4.2, 108.0), "Marrocos": (31.8, -7.1), "Mauritânia": (21.0, -10.9),
    "Maurício": (-20.3, 57.5), "Mayotte": (-12.8, 45.2), "Moldova": (47.4, 28.4),
    "Moçambique": (-18.7, 35.5), "Myanmar": (17.1, 96.1), "México": (23.6, -102.6),
    "Namíbia": (-22.0, 17.1), "Nicarágua": (12.9, -85.2), "Nigéria": (9.1, 8.7),
    "Noruega": (60.5, 8.5), "Nova Zelândia": (-41.0, 174.9), "Níger": (17.6, 8.1),
    "Omã": (21.5, 55.9), "Palestina": (31.9, 35.2), "Panamá": (8.5, -80.8),
    "Papua Nova Guiné": (-6.3, 143.9), "Paquistão": (30.4, 69.3), "Paraguai": (-23.4, -58.4),
    "Peru": (-9.2, -75.0), "Polônia": (51.9, 19.1), "Portugal": (39.4, -8.2),
    "Qatar": (25.4, 51.2), "Quênia": (0.0, 37.9), "RD Congo": (-4.0, 21.8),
    "Reino Unido": (54.4, -3.4), "República Dominicana": (18.7, -70.2), "República Tcheca": (49.8, 15.5),
    "Reunião": (-21.1, 55.5), "Romênia": (45.9, 24.9), "Ruanda": (-1.9, 29.9),
    "Samoa": (-13.8, -172.1), "Senegal": (14.5, -14.5), "Serra Leoa": (8.5, -11.8),
    "Singapura": (1.4, 103.8), "Somaliland": (10.5, 46.2), "Somália": (5.2, 46.2),
    "Sri Lanka": (7.9, 80.8), "Sudão": (12.9, 30.2), "Sudão do Sul": (4.9, 31.3),
    "Suriname": (3.9, -56.0), "Suécia": (60.1, 18.6), "Suíça": (46.8, 8.2),
    "Sérvia": (44.0, 21.0), "Taiwan": (23.7, 121.0), "Tanzânia": (-6.4, 34.9),
    "Togo": (8.6, 0.8), "Trinidad e Tobago": (10.7, -61.2), "Tunísia": (33.9, 9.6),
    "Turquia": (38.9, 35.2), "Ucrânia": (48.4, 31.2), "Uganda": (1.4, 32.3),
    "Uruguai": (-32.5, -55.8), "Uzbequistão": (41.4, 64.6), "Venezuela": (6.4, -66.6),
    "Vietnã": (14.1, 108.3), "Zimbábue": (-19.0, 29.2), "Zâmbia": (-13.1, 27.9),
    "Áustria": (47.5, 14.6),
}

GEO_REGIONS: dict[tuple[str,str], tuple[float,float]] = {
    # Africa do Sul
    ("Africa do Sul","Gauteng"): (-26.1,28.0), ("Africa do Sul","Western Cape"): (-33.9,18.4),
    ("Africa do Sul","KwaZulu-Natal"): (-29.8,31.0), ("Africa do Sul","Eastern Cape"): (-31.5,27.5),
    ("Africa do Sul","Other Provinces"): (-26.0,26.0),
    # Alemanha
    ("Alemanha","Baden-Württemberg"): (48.7,9.2), ("Alemanha","Bavaria"): (48.1,11.6),
    ("Alemanha","Berlin"): (52.5,13.4), ("Alemanha","North Rhine-Westphalia"): (51.5,7.5),
    ("Alemanha","Other States"): (52.0,11.0),
    # Argentina
    ("Argentina","Centro (Buenos Aires)"): (-34.6,-58.4), ("Argentina","Noroeste"): (-24.0,-65.4),
    ("Argentina","Nordeste"): (-27.5,-57.0), ("Argentina","Cuyo"): (-32.9,-68.8),
    ("Argentina","Patagônia"): (-41.8,-65.0), ("Argentina","Other"): (-37.0,-61.0),
    # Arábia Saudita
    ("Arábia Saudita","Riyadh"): (24.7,46.7), ("Arábia Saudita","Makkah (Jeddah)"): (21.4,39.8),
    ("Arábia Saudita","Eastern Province"): (26.4,50.0), ("Arábia Saudita","Al Madinah"): (24.5,39.6),
    ("Arábia Saudita","Other Regions"): (22.0,44.0),
    # Austrália
    ("Austrália","New South Wales"): (-33.9,151.2), ("Austrália","Victoria"): (-37.8,145.0),
    ("Austrália","Queensland"): (-27.5,153.0), ("Austrália","Western Australia"): (-31.9,115.9),
    ("Austrália","South Australia"): (-34.9,138.6), ("Austrália","Other"): (-12.5,130.8),
    # Bangladesh
    ("Bangladesh","Dhaka"): (23.8,90.4), ("Bangladesh","Chittagong"): (22.3,91.8),
    ("Bangladesh","Rajshahi"): (24.4,88.6), ("Bangladesh","Khulna"): (22.8,89.6),
    ("Bangladesh","Sylhet"): (24.9,91.9), ("Bangladesh","Other"): (25.0,89.4),
    # Canadá
    ("Canadá","Atlantic Canada"): (46.0,-64.0), ("Canadá","Quebec"): (52.0,-72.0),
    ("Canadá","Ontario"): (50.0,-86.0), ("Canadá","Prairie Provinces"): (52.0,-106.0),
    ("Canadá","British Columbia"): (53.7,-127.6),
    # China
    ("China","Norte (Beijing/Hebei)"): (40.0,116.4), ("China","Nordeste (Manchúria)"): (44.0,126.0),
    ("China","Leste (Shanghai/Jiangsu)"): (31.2,121.5), ("China","Sul (Guangdong)"): (23.1,113.3),
    ("China","Centro (Hubei/Hunan)"): (30.6,114.3), ("China","Sudoeste (Sichuan/Yunnan)"): (30.7,104.1),
    ("China","Noroeste"): (36.1,103.8), ("China","Other"): (35.0,105.0),
    # Colômbia
    ("Colômbia","Caribe"): (10.9,-74.8), ("Colômbia","Andina"): (4.7,-74.1),
    ("Colômbia","Pacífico"): (3.9,-77.0), ("Colômbia","Orinoquía"): (4.2,-72.5),
    ("Colômbia","Amazônica"): (-1.5,-71.5), ("Colômbia","Nororiental"): (7.1,-73.1),
    ("Colômbia","Other"): (6.2,-75.6),
    # Coreia do Sul
    ("Coreia do Sul","Capital Region (Seoul)"): (37.6,126.9), ("Coreia do Sul","Gyeongsang"): (35.2,128.7),
    ("Coreia do Sul","Jeolla"): (35.8,127.1), ("Coreia do Sul","Chungcheong"): (36.5,127.7),
    ("Coreia do Sul","Other"): (37.9,127.7),
    # EAU
    ("EAU","Dubai"): (25.2,55.3), ("EAU","Abu Dhabi"): (24.5,54.4),
    ("EAU","Sharjah & Northern Emirates"): (25.3,55.4),
    # EUA
    ("EUA","Northeast"): (42.0,-73.0), ("EUA","Midwest"): (41.0,-87.0),
    ("EUA","South"): (33.0,-84.0), ("EUA","West"): (37.0,-120.0),
    # Egito
    ("Egito","Cairo & Giza"): (30.0,31.2), ("Egito","Delta do Nilo"): (30.8,31.0),
    ("Egito","Upper Egypt"): (26.0,32.0), ("Egito","Canal & Sinai"): (30.5,32.5),
    # Espanha
    ("Espanha","Catalunha"): (41.4,2.2), ("Espanha","Madrid"): (40.4,-3.7),
    ("Espanha","Andaluzia"): (37.3,-6.0), ("Espanha","Comunidade Valenciana"): (39.5,-0.5),
    ("Espanha","Other Regions"): (42.0,-2.0),
    # França
    ("França","Île-de-France"): (48.9,2.3), ("França","Auvergne-Rhône-Alpes"): (45.7,4.8),
    ("França","Provence-Alpes-Côte d'Azur"): (43.9,6.1), ("França","Occitanie"): (43.6,3.9),
    ("França","Other Regions"): (47.0,1.0),
    # Grécia
    ("Grécia","Attica (Athens)"): (37.9,23.7), ("Grécia","Central Greece"): (38.8,22.4),
    ("Grécia","North Greece (Thessaloniki)"): (40.6,22.9), ("Grécia","Peloponnese & Crete"): (37.2,22.3),
    ("Grécia","Islands"): (37.4,25.0),
    # Indonésia
    ("Indonésia","Jawa"): (-6.9,107.6), ("Indonésia","Sumatera"): (0.6,101.3),
    ("Indonésia","Kalimantan"): (1.7,113.9), ("Indonésia","Sulawesi"): (-1.4,121.4),
    ("Indonésia","Bali e Nusa Tenggara"): (-8.5,115.2), ("Indonésia","Papua e Maluku"): (-4.3,138.0),
    # Iraque
    ("Iraque","Baghdad"): (33.3,44.4), ("Iraque","South Iraq"): (31.0,46.2),
    ("Iraque","Kurdistan Region"): (36.2,44.0), ("Iraque","North & West Iraq"): (34.6,42.8),
    # Itália
    ("Itália","Nord-Ovest (Milano)"): (45.5,9.2), ("Itália","Nord-Est (Venezia)"): (45.4,12.3),
    ("Itália","Centro (Roma)"): (41.9,12.5), ("Itália","Sud (Naples)"): (40.8,14.3),
    ("Itália","Isole (Sicília/Sardenha)"): (37.5,14.0),
    # Japão
    ("Japão","Kantō (Tokyo)"): (35.7,139.7), ("Japão","Kansai (Osaka)"): (34.7,135.5),
    ("Japão","Tōkai (Nagoya)"): (35.2,137.0), ("Japão","Tōhoku"): (38.3,140.9),
    ("Japão","Kyushu e Okinawa"): (33.6,130.4), ("Japão","Other"): (43.1,141.4),
    # Malásia
    ("Malásia","Peninsular Malaysia"): (3.8,103.1), ("Malásia","Sabah"): (5.0,117.1),
    ("Malásia","Sarawak"): (2.5,113.1),
    # Marrocos
    ("Marrocos","Grand Casablanca-Settat"): (33.6,-7.6), ("Marrocos","Rabat-Salé"): (34.0,-6.8),
    ("Marrocos","Fès-Meknès"): (34.0,-5.0), ("Marrocos","Marrakech-Safi"): (31.6,-8.0),
    ("Marrocos","Other Regions"): (31.0,-5.0),
    # México
    ("México","Norte"): (27.0,-107.0), ("México","Centro-Norte"): (24.0,-104.0),
    ("México","Centro"): (19.4,-99.1), ("México","Sur-Sureste"): (18.0,-92.0),
    ("México","Peninsular"): (20.7,-88.0),
    # Nigéria
    ("Nigéria","South West (Lagos)"): (6.5,3.4), ("Nigéria","South South (Rivers)"): (4.8,7.0),
    ("Nigéria","South East"): (5.5,8.0), ("Nigéria","North West"): (12.0,8.0),
    ("Nigéria","North Central"): (9.1,7.5), ("Nigéria","North East"): (11.0,13.1),
    # Paquistão
    ("Paquistão","Punjab"): (31.5,74.3), ("Paquistão","Sindh"): (25.9,68.3),
    ("Paquistão","Khyber Pakhtunkhwa"): (34.2,71.9), ("Paquistão","Balochistan"): (28.5,65.5),
    ("Paquistão","Islamabad"): (33.7,73.1),
    # Polônia
    ("Polônia","Mazowieckie (Warszawa)"): (52.2,21.0), ("Polônia","Śląskie"): (50.3,19.0),
    ("Polônia","Małopolskie"): (50.1,20.0), ("Polônia","Wielkopolskie"): (52.4,17.0),
    ("Polônia","Other Provinces"): (52.0,18.5),
    # Portugal
    ("Portugal","Lisboa e Vale do Tejo"): (38.7,-9.1), ("Portugal","Norte"): (41.5,-8.4),
    ("Portugal","Centro"): (39.8,-8.0), ("Portugal","Alentejo"): (38.0,-7.9),
    ("Portugal","Algarve"): (37.1,-8.1),
    # Quênia
    ("Quênia","Nairobi"): (-1.3,36.8), ("Quênia","Central"): (-0.5,37.1),
    ("Quênia","Coast"): (-3.2,40.1), ("Quênia","Nyanza & Western"): (-0.1,34.8),
    ("Quênia","Rift Valley"): (0.5,36.0), ("Quênia","Eastern & NEP"): (0.7,38.5),
    # Reino Unido
    ("Reino Unido","England"): (52.4,-1.8), ("Reino Unido","Scotland"): (56.5,-4.2),
    ("Reino Unido","Wales"): (52.1,-3.8), ("Reino Unido","Northern Ireland"): (54.6,-6.7),
    # Romênia
    ("Romênia","Bucharest-Ilfov"): (44.4,26.1), ("Romênia","Sud-Muntenia"): (44.1,25.4),
    ("Romênia","Nord-Vest"): (46.7,23.6), ("Romênia","Nord-Est"): (47.2,26.7),
    ("Romênia","Other Regions"): (45.8,24.6),
    # Tanzânia
    ("Tanzânia","Dar es Salaam"): (-6.8,39.3), ("Tanzânia","Zone Kaskazini (Northern)"): (-3.4,36.7),
    ("Tanzânia","Zone Kanda ya Kati (Central)"): (-5.8,35.7), ("Tanzânia","Zone Kusini (Southern)"): (-9.3,35.7),
    ("Tanzânia","Zone Mashariki (Eastern)"): (-6.2,38.5), ("Tanzânia","Zone Magharibi (Western)"): (-4.9,32.8),
    # Turquia
    ("Turquia","Marmara (Istanbul)"): (41.0,28.9), ("Turquia","Ege (Izmir)"): (38.4,27.1),
    ("Turquia","Akdeniz"): (37.0,35.3), ("Turquia","İç Anadolu (Ankara)"): (39.9,32.9),
    ("Turquia","Karadeniz"): (41.3,36.2), ("Turquia","Doğu Anadolu"): (38.8,40.5),
    ("Turquia","Güneydoğu Anadolu"): (37.2,40.1),
    # Ucrânia
    ("Ucrânia","Eastern Ukraine (Kharkiv/Dnipro)"): (49.0,36.2), ("Ucrânia","Central (Kyiv)"): (50.5,30.5),
    ("Ucrânia","Southern (Odessa)"): (46.5,30.7), ("Ucrânia","Western Ukraine"): (49.8,24.0),
    ("Ucrânia","Northern Ukraine"): (51.5,31.3), ("Ucrânia","Other"): (49.0,33.0),
    # Vietnã
    ("Vietnã","Red River Delta (Hanoi)"): (21.0,105.9), ("Vietnã","Southeast (Ho Chi Minh)"): (10.8,106.7),
    ("Vietnã","Mekong River Delta"): (10.4,105.7), ("Vietnã","North Central & Central Coast"): (16.5,107.6),
    ("Vietnã","Other Regions"): (15.0,107.0),
    # Argélia
    ("Argélia","Norte (Alger)"): (36.7,3.1), ("Argélia","Nordeste"): (36.4,6.6),
    ("Argélia","Noroeste"): (35.7,1.3), ("Argélia","Sul (Saara)"): (27.0,2.9),
}

def get_geo(loc_key, ciclo):
    """Return geo dict for a BDOS loc_key. Handles Brazil (GEO_L5) and Chile (CL{n} - CITY)."""
    if loc_key in GEO_L5:
        return GEO_L5[loc_key]
    if ciclo == "BAT CHILE":
        m = re.match(r"CL\d+ - (.+)$", loc_key)
        if m:
            city = m.group(1).strip()
            g = GEO_CHILE_REGIONS.get(city)
            if g:
                return {"lat": g["lat"], "lng": g["lng"],
                        "city": city.title(), "state": "", "region": city, "country": "Chile"}
    return None

def parse_num(s):
    if s is None or s == "" or s == "-": return None
    if isinstance(s, (int, float)):
        v = float(s)
        # Round to integer if very close (handles Excel floating-point formula noise)
        return round(v) if abs(v - round(v)) < 0.5 else round(v, 4)
    s = str(s).replace("$","").replace(".","").replace(",",".").strip()
    s = re.sub(r"[^\d\.\-]","",s)
    try: return float(s)
    except: return None

def parse_pct(s):
    if s is None or s == "" or s == "-": return None
    if isinstance(s, (int, float)):
        v = float(s)
        return round(v * 100, 4) if v < 1.5 else v  # normalize decimal fractions
    s = str(s).replace("%","").replace(",",".").strip()
    try:
        v = float(s)
        return round(v * 100, 4) if v < 1.5 else v
    except: return None

# ─── READ MARKET INTELLIGENCE ─────────────────────────────────────────────────
print("Reading Market Intelligence...")
wb_mi = load_workbook(F2, read_only=True, data_only=True)

# By Country (rows 5+, row 4 = header)
ws_country = wb_mi["🌍 By Country"]
rows_country = list(ws_country.iter_rows(min_row=4, values_only=True))
hdr_c = [str(c) if c else "" for c in rows_country[0]]

countries_data = {}
for row in rows_country[1:]:
    if not row[1] or not row[2]: continue
    subregion = str(row[1]).strip()
    country   = str(row[2]).strip()
    if not country or country.startswith("BAT"): continue
    countries_data[country] = {
        "country":     country,
        "subregion":   subregion,
        "continent":   CONTINENT_MAP.get(subregion, "Other"),
        "gdp":         parse_num(row[3]),
        "population":  parse_num(row[4]),
        "income_pc":   parse_num(row[5]),
        "smokers_est": parse_num(row[6]),
        "smokers_pct": parse_pct(row[7]),
        "bat_share":   parse_pct(row[8]),
        "tobacco_mkt": parse_num(row[9]),
        "pos_total":   parse_num(row[10]),
        "pos_bat":     parse_num(row[11]),
        "labour_cost": parse_num(row[12]),
        "pos_active":  parse_num(row[13]),
        "pos_active_pct": parse_pct(row[14]),
        "posms":       parse_num(row[15]),
        "posms_pct":   parse_pct(row[16]),
        "op_cost":     parse_num(row[17]),
        "data_type":   "estimated",
        "regions":     {}
    }

# By Region (rows 5+, row 4 = header)
ws_region = wb_mi["📍 By Region"]
rows_region = list(ws_region.iter_rows(min_row=4, values_only=True))

for row in rows_region[1:]:
    if not row[1] or not row[2] or not row[3]: continue
    country = str(row[2]).strip()
    region  = str(row[3]).strip()
    if country not in countries_data: continue
    if country == "Chile":
        geo_r = GEO_CHILE_REGIONS.get(region.upper(), {})
        lat, lng = geo_r.get("lat"), geo_r.get("lng")
    else:
        geo_key = GEO_REGIONS.get((country, region))
        if not geo_key:
            # fallback: country centroid (covers "Nacional (X)" regions)
            centroid = COUNTRY_CENTROIDS.get(country)
            geo_key = centroid
        lat = geo_key[0] if geo_key else None
        lng = geo_key[1] if geo_key else None
    countries_data[country]["regions"][region] = {
        "region":      region,
        "country":     country,
        "gdp":         parse_num(row[4]),
        "population":  parse_num(row[5]),
        "income_pc":   parse_num(row[6]),
        "smokers_est": parse_num(row[7]),
        "smokers_pct": parse_pct(row[8]),
        "bat_share":   parse_pct(row[9]),
        "tobacco_mkt": parse_num(row[10]),
        "pos_total":   parse_num(row[11]),
        "pos_bat":     parse_num(row[12]),
        "labour_cost": parse_num(row[13]),
        "pos_active":  parse_num(row[14]),
        "pos_active_pct": parse_pct(row[15]),
        "posms":       parse_num(row[16]),
        "posms_pct":   parse_pct(row[17]),
        "lat":         lat,
        "lng":         lng,
        "data_type":   "estimated",
        "cities":      {}
    }

wb_mi.close()
print(f"  Countries loaded: {len(countries_data)}")
print(f"  Brazil regions: {list(countries_data.get('Brazil',{}).get('regions',{}).keys())}")
print(f"  Chile regions: {list(countries_data.get('Chile',{}).get('regions',{}).keys())[:5]}... ({len(countries_data.get('Chile',{}).get('regions',{}))} total)")

# ─── READ BDOS (operational data) ─────────────────────────────────────────────
print("\nReading BDOS (sampling for correlation analysis)...")
wb_ops = load_workbook(F1, read_only=True, data_only=True)
ws_bdos = wb_ops["BDOS"]

rows_iter = ws_bdos.iter_rows(min_row=1, values_only=True)
hdr_b = [str(c) if c else "" for c in next(rows_iter)]

# Col indices (0-based)
COL = {h: i for i, h in enumerate(hdr_b)}
print(f"  BDOS columns: {list(COL.keys())[:10]}...")

# Aggregate by location key
loc_stats = {}   # loc_key → {os_count, productive, sla_total, sla_count, rating_total, rating_count, ciclo, statuses}
ciclo_counts = {}

row_count = 0
for row in rows_iter:
    row_count += 1
    ciclo = str(row[COL.get("Nome Ciclo", 2)] or "").strip()
    loc   = str(row[COL.get("Descrição (Local)", 13)] or "").strip()
    status = str(row[COL.get("Status", 9)] or "").strip()
    prod  = str(row[COL.get("Produtiva", 23)] or "").strip()
    sla   = row[COL.get("Dias para SLA", 22)]
    rating= row[COL.get("Avaliação da OS", 27)]
    camp  = str(row[COL.get("Descrição (Campanhas)", 26)] or "").strip()

    if ciclo: ciclo_counts[ciclo] = ciclo_counts.get(ciclo, 0) + 1

    if not loc or loc == "<SEM LOCAL>": continue
    if loc not in loc_stats:
        loc_stats[loc] = {"os":0,"prod":0,"sla_sum":0,"sla_n":0,"rate_sum":0,"rate_n":0,
                          "ciclo":ciclo,"statuses":{},"campaigns":{}}
    s = loc_stats[loc]
    s["os"] += 1
    if prod == "Sim": s["prod"] += 1
    if sla is not None:
        try: s["sla_sum"] += float(str(sla).replace(",",".")); s["sla_n"] += 1
        except: pass
    if rating is not None and str(rating).strip() not in ("","None"):
        try: s["rate_sum"] += float(str(rating).replace(",",".")); s["rate_n"] += 1
        except: pass
    s["statuses"][status] = s["statuses"].get(status, 0) + 1
    if camp and camp not in ("NENHUMA",""):
        s["campaigns"][camp] = s["campaigns"].get(camp, 0) + 1

wb_ops.close()
print(f"  Total rows processed: {row_count:,}")
print(f"  Ciclos (países): {ciclo_counts}")
print(f"  Unique locations: {len(loc_stats)}")

# ─── CORRELATION REPORT ───────────────────────────────────────────────────────
print("\n" + "="*70)
print("CORRELATION ANALYSIS — BRAZIL & CHILE")
print("="*70)

print("\n── JOIN KEY: BDOS.Nome Ciclo → Market Intel Country ──")
for ciclo, count in sorted(ciclo_counts.items(), key=lambda x: -x[1]):
    mapped = CICLO_TO_COUNTRY.get(ciclo, "⚠ UNMAPPED")
    in_mi  = "✓ IN MI" if mapped in countries_data else "✗ NOT IN MI"
    print(f"  BDOS[{ciclo}] ({count:,} OSs) → MI[{mapped}] {in_mi}")

print("\n── BDOS Locations → Region Mapping ──")
join_ok = []; join_fail = []
for loc_key, stats in sorted(loc_stats.items()):
    ciclo  = stats["ciclo"]
    geo    = get_geo(loc_key, ciclo)
    prefix = loc_key.split(" - ")[0].strip()
    if CICLO_TO_COUNTRY.get(ciclo) == "Chile":
        mi_reg = geo.get("region") if geo else None
    else:
        mi_reg = BDOS_REGION_MAP.get(prefix)
    country= CICLO_TO_COUNTRY.get(ciclo,"?")
    in_mi_reg = (country in countries_data and
                 mi_reg and mi_reg in countries_data.get(country,{}).get("regions",{}))
    pct_prod = round(stats["prod"]/stats["os"]*100,1) if stats["os"] else 0
    avg_sla  = round(stats["sla_sum"]/stats["sla_n"],1) if stats["sla_n"] else None
    avg_rate = round(stats["rate_sum"]/stats["rate_n"],1) if stats["rate_n"] else None
    has_geo  = "✓geo" if geo else "✗geo"
    reg_ok   = "✓reg" if in_mi_reg else "✗reg"
    print(f"  {has_geo} {reg_ok} [{loc_key}]")
    print(f"        country={country} | mi_region={mi_reg} | OSs={stats['os']:,} | prod={pct_prod}% | SLA={avg_sla}d | rating={avg_rate}")
    if geo:
        print(f"        WGS84: lat={geo['lat']}, lng={geo['lng']} | city={geo['city']}/{geo['state']}")
    if not geo: join_fail.append(loc_key)
    else: join_ok.append(loc_key)

print(f"\n── SUMMARY ──")
print(f"  Locations WITH geocoding: {len(join_ok)}")
print(f"  Locations WITHOUT geocoding: {len(join_fail)}")
if join_fail:
    print(f"  UNMAPPED: {join_fail}")

# ─── BUILD HIERARCHICAL JSON ──────────────────────────────────────────────────
print("\nBuilding hierarchical JSON L0→L5...")

# Inject operational data into Brazil regions (L4) and cities (L5)
brazil = countries_data.get("Brazil", {})
brazil["data_type"] = "real+estimated"

for loc_key, stats in loc_stats.items():
    ciclo   = stats["ciclo"]
    geo = get_geo(loc_key, ciclo)
    if not geo: continue
    country = CICLO_TO_COUNTRY.get(ciclo)
    if not country or country not in countries_data: continue

    prefix  = loc_key.split(" - ")[0].strip()
    if country == "Chile":
        # For Chile, region key == city name extracted from "CL{n} - CITY"
        mi_reg = geo.get("region")
    else:
        mi_reg = BDOS_REGION_MAP.get(prefix)
    if not mi_reg: continue

    c_data = countries_data[country]
    if mi_reg not in c_data["regions"]: continue

    reg_data = c_data["regions"][mi_reg]
    if "cities" not in reg_data: reg_data["cities"] = {}

    city_key = f"{geo['city']}/{geo['state']}"
    pct_prod = round(stats["prod"]/stats["os"]*100,1) if stats["os"] else 0
    avg_sla  = round(stats["sla_sum"]/stats["sla_n"],1) if stats["sla_n"] else None
    avg_rate = round(stats["rate_sum"]/stats["rate_n"],1) if stats["rate_n"] else None

    reg_data["cities"][city_key] = {
        "city":           geo["city"],
        "state":          geo["state"],
        "lat":            geo["lat"],
        "lng":            geo["lng"],
        "loc_key":        loc_key,
        "country":        country,
        "region":         mi_reg,
        "os_total":       stats["os"],
        "os_productive":  stats["prod"],
        "productive_pct": pct_prod,
        "sla_avg_days":   avg_sla,
        "avg_rating":     avg_rate,
        "top_campaigns":  dict(sorted(stats["campaigns"].items(), key=lambda x:-x[1])[:10]),
        "statuses":       stats["statuses"],
        "data_type":      "real"
    }
    # Accumulate operational totals into region
    reg_data.setdefault("os_total", 0)
    reg_data.setdefault("os_productive", 0)
    reg_data["os_total"]      += stats["os"]
    reg_data["os_productive"] += stats["prod"]
    reg_data["data_type"] = "real+estimated"
    c_data["data_type"]   = "real+estimated"

# Finalize region productive_pct
for country, cd in countries_data.items():
    for reg, rd in cd.get("regions",{}).items():
        if rd.get("os_total",0) > 0:
            rd["productive_pct"] = round(rd["os_productive"]/rd["os_total"]*100,1)

# Build full L0→L5 tree
hierarchy = {"level": 0, "name": "Mundo", "continents": {}}
for l1 in L1_ORDER:
    hierarchy["continents"][l1] = {"level":1,"name":l1,"subregions":{}}

for country, cd in countries_data.items():
    l1 = cd["continent"]
    l2 = cd["subregion"]
    if l1 not in hierarchy["continents"]: continue
    cont = hierarchy["continents"][l1]
    if l2 not in cont["subregions"]:
        cont["subregions"][l2] = {"level":2,"name":l2,"continent":l1,"countries":{}}
    cont["subregions"][l2]["countries"][country] = cd

# Save outputs
(OUT/"hierarchy.json").write_text(
    json.dumps(hierarchy, ensure_ascii=False, indent=2), encoding="utf-8")

(OUT/"countries.json").write_text(
    json.dumps(countries_data, ensure_ascii=False, indent=2), encoding="utf-8")

(OUT/"geocoding_l5.json").write_text(
    json.dumps(GEO_L5, ensure_ascii=False, indent=2), encoding="utf-8")

# Summary stats
all_cities = sum(len(cd.get("regions",{}).get(r,{}).get("cities",{}))
    for cd in countries_data.values() for r in cd.get("regions",{}))
total_os = sum(s["os"] for s in loc_stats.values())

print(f"\n── OUTPUT FILES ──")
print(f"  {OUT/'hierarchy.json'} ({(OUT/'hierarchy.json').stat().st_size//1024}KB)")
print(f"  {OUT/'countries.json'} ({(OUT/'countries.json').stat().st_size//1024}KB)")
print(f"  {OUT/'geocoding_l5.json'}")
print(f"\n── FINAL STATS ──")
print(f"  L1 continents:    {len(hierarchy['continents'])}")
print(f"  L2 sub-regions:   {sum(len(c['subregions']) for c in hierarchy['continents'].values())}")
print(f"  L3 countries:     {len(countries_data)}")
print(f"  L4 regions:       {sum(len(cd.get('regions',{})) for cd in countries_data.values())}")
print(f"  L5 cities (real): {all_cities}")
print(f"  Total OSs mapped: {total_os:,}")

# Download world GeoJSON for choropleth map
import urllib.request
world_geojson_path = OUT / "world.geojson"
if not world_geojson_path.exists():
    print("\nDownloading world GeoJSON...")
    url = "https://raw.githubusercontent.com/holtzy/D3-graph-gallery/master/DATA/world.geojson"
    try:
        urllib.request.urlretrieve(url, world_geojson_path)
        print(f"  Downloaded world.geojson ({world_geojson_path.stat().st_size//1024}KB)")
    except Exception as e:
        print(f"  Primary URL failed: {e}, trying fallback...")
        try:
            url2 = "https://raw.githubusercontent.com/nvkelso/natural-earth-vector/master/geojson/ne_110m_admin_0_countries.geojson"
            urllib.request.urlretrieve(url2, world_geojson_path)
            print(f"  Downloaded world.geojson from fallback ({world_geojson_path.stat().st_size//1024}KB)")
        except Exception as e2:
            print(f"  Fallback also failed: {e2}")
else:
    print(f"\n  world.geojson already exists ({world_geojson_path.stat().st_size//1024}KB)")

print("\nDone.")
